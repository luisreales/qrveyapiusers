from io import BytesIO
from openpyxl import Workbook
from flask import Response


def generate_excel(title, data, filename):
    wb = Workbook()
    ws = wb.active

  
    for i, t in enumerate(title):
        ws.cell(row=1, column=(i + 1)).value = t[1]
 
    title_fields = [t[0] for t in title]
    for i, _data in enumerate(data):
        one_row = [_data[t] for t in title_fields]
        for j, d in enumerate(one_row):
            ws.cell(row=(i + 2), column=(j + 1)).value = d

    
    sio = BytesIO()
    wb.save(sio)

    response = Response()
    response.headers.add("Content-Type", "application/vnd.ms-excel")
    response.headers.add('Content-Disposition', 'attachment', filename=filename.encode("utf-8").decode("latin1"))
    sio.seek(0)
    response.data = sio.getvalue()
    return response